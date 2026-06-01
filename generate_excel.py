import sys, json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import copy

MONTHS_FR = ["Janvier","Février","Mars","Avril","Mai","Juin",
             "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

def generate(entries_json, output_path):
    entries = json.loads(entries_json)
    wb = load_workbook('/home/claude/ndf-app/template.xlsx')
    ws = wb['NDF']

    # Detect month/year from first entry
    if entries:
        first_date = datetime.strptime(entries[0]['date'], '%Y-%m-%d')
        mois = MONTHS_FR[first_date.month - 1]
        annee = str(first_date.year)
    else:
        now = datetime.now()
        mois = MONTHS_FR[now.month - 1]
        annee = str(now.year)

    # Replace XmoisX and XannéeX
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'XmoisX':
                cell.value = mois
            elif cell.value == 'XannéeX':
                cell.value = annee

    # Fill data rows 15 to 41
    for i, entry in enumerate(entries[:27]):
        row = 15 + i
        date_obj = datetime.strptime(entry['date'], '%Y-%m-%d')
        ws.cell(row=row, column=1).value = date_obj.strftime('%d/%m/%Y')
        ws.cell(row=row, column=2).value = entry.get('libelle', '')
        ws.cell(row=row, column=3).value = entry.get('kms', None)
        ws.cell(row=row, column=4).value = entry.get('repas', None)
        ws.cell(row=row, column=5).value = entry.get('hotel', None)
        ws.cell(row=row, column=6).value = entry.get('taxis', None)
        ws.cell(row=row, column=7).value = entry.get('divers', None)

    wb.save(output_path)
    return {'status': 'ok', 'mois': mois, 'annee': annee, 'nb_entries': len(entries)}

if __name__ == '__main__':
    result = generate(sys.argv[1], sys.argv[2])
    print(json.dumps(result))
